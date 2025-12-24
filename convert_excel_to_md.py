import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys

def excel_to_markdown(excel_path, md_path):
    wb = openpyxl.load_workbook(excel_path)
    
    with open(md_path, 'w', encoding='utf-8') as f:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            f.write(f'# {sheet_name}\n\n')
            
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            if not data:
                f.write('*空表*\n\n')
                continue
            
            max_cols = max(len(row) for row in data)
            
            for i, row in enumerate(data):
                row = list(row) + [''] * (max_cols - len(row))
                
                if i == 0:
                    f.write('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in row) + ' |\n')
                    f.write('| ' + ' | '.join(['---'] * max_cols) + ' |\n')
                else:
                    f.write('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in row) + ' |\n')
            
            f.write('\n---\n\n')
    
    wb.close()

if __name__ == '__main__':
    if len(sys.argv) == 3:
        excel_path = sys.argv[1]
        md_path = sys.argv[2]
        if os.path.exists(excel_path):
            excel_to_markdown(excel_path, md_path)
            print(f'已转换: {excel_path} -> {md_path}')
        else:
            print(f'文件不存在: {excel_path}')
    else:
        print('用法: python convert_excel_to_md.py <excel文件路径> <md文件路径>')
